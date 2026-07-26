"""
seed_projections.py
────────────────────
Seeds player_weekly_projections (season 2026, week 1, source='internal_model')
from the "Projected Pts (PPR)" column already present in oneleague_prices1.xlsx
— the same file migrate_to_sportsdata.py uses for prices, reusing its name
aliases. Season-long PPR projection is divided evenly across 17 weeks.

Usage:
  python3 seed_projections.py
"""

import openpyxl
import mysql.connector
import os
from decimal import Decimal

env_path = os.path.join(os.path.dirname(__file__), '.env')
env = {}
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if line and '=' in line and not line.startswith('#'):
            k, _, v = line.partition('=')
            env[k.strip()] = v.strip()

DB = dict(
    host=env['MYSQL_HOST'],
    port=int(env.get('MYSQL_PORT', 3306)),
    user=env['MYSQL_USER'],
    password=env['MYSQL_PASSWORD'],
    database=env['MYSQL_DATABASE'],
)

SEASON_YEAR = 2026
WEEKS       = 17
SOURCE      = 'internal_model'

NAME_ALIASES = {
    'Cam Ward':              'Cameron Ward',
    'Stetson Bennett':       'Stetson Bennett IV',
    'Joe Milton':            'Joe Milton III',
    'Travis Etienne':        'Travis Etienne Jr.',
    'Aaron Jones':           'Aaron Jones Sr.',
    'Brian Robinson':        'Brian Robinson Jr.',
    'Tyrone Tracy':          'Tyrone Tracy Jr.',
    'Chris Rodriguez':       'Chris Rodriguez Jr.',
    'LeQuint Allen':         'LeQuint Allen Jr.',
    'Ollie Gordon':          'Ollie Gordon II',
    'Michael Pittman':       'Michael Pittman Jr.',
    'D.J. Moore':            'DJ Moore',
    'Luther Burden':         'Luther Burden III',
    'Chris Godwin':          'Chris Godwin Jr.',
    'Tre Harris':            "Tre' Harris",
    'Demario Douglas':       'DeMario Douglas',
    'Theo Wease':            'Theo Wease Jr.',
    'Marvin Mims':           'Marvin Mims Jr.',
    'KeAndre Lambert-Smith': 'Keandre Lambert-Smith',
    'Kevin Austin':          'Kevin Austin Jr.',
    "Dont'e Thornton":       "Dont'e Thornton Jr.",
    'Terrace Marshall':      'Terrace Marshall Jr.',
    'Jimmy Horn':            'Jimmy Horn Jr.',
    'Efton Chism':           'Efton Chism III',
    'Harold Fannin':         'Harold Fannin Jr.',
    'Kyle Pitts':            'Kyle Pitts Sr.',
    'Chigoziem Okonkwo':     'Chig Okonkwo',
    'Erick All':             'Erick All Jr.',
    'Andrew Ogletree':       'Drew Ogletree',
}

print('Loading spreadsheet projections...')
xlsx = os.path.join(os.path.dirname(__file__), 'oneleague_prices1.xlsx')
wb   = openpyxl.load_workbook(xlsx, data_only=True)

sheet_proj = {}  # sdio_name -> Decimal season-long PPR projection
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name, proj = row[0], row[4]
        if not name or proj is None:
            continue
        sdio_name = NAME_ALIASES.get(name.strip(), name.strip())
        sheet_proj[sdio_name] = Decimal(str(proj))

print(f'  {len(sheet_proj)} projection entries loaded')

conn = mysql.connector.connect(**DB)
cur  = conn.cursor()

cur.execute('SELECT id, full_name FROM players')
db_players = cur.fetchall()

rows = []
matched = 0
for player_id, full_name in db_players:
    season_proj = sheet_proj.get(full_name)
    if season_proj is None:
        continue
    matched += 1
    weekly = (season_proj / WEEKS).quantize(Decimal('0.01'))
    for week in range(1, WEEKS + 1):
        rows.append((player_id, SEASON_YEAR, week, SOURCE, weekly))

print(f'Matched {matched} players, inserting {len(rows)} weekly projection rows...')

cur.executemany("""
    INSERT INTO player_weekly_projections
        (player_id, season_year, week, projection_source, expected_points)
    VALUES (%s, %s, %s, %s, %s)
    ON DUPLICATE KEY UPDATE expected_points = VALUES(expected_points)
""", rows)
conn.commit()
print(f'Done. {cur.rowcount} rows affected.')

cur.close()
conn.close()
