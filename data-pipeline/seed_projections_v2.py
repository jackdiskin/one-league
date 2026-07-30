"""
seed_projections_v2.py
────────────────────────
Seeds player_weekly_projections (season 2026, weeks 1-17, source='internal_model')
from the "Projected Pts (PPR)" column in oneleague_prices2.xlsx — the same
sheet/matching approach as set_2026_prices_v2.py (name-based, same alias
table), so projection coverage now matches price coverage exactly, including
every newly-backfilled rookie that's present in the sheet. Season-long PPR
projection is divided evenly across 17 weeks.

Upserts (ON DUPLICATE KEY UPDATE), so re-running after another spreadsheet
edit is safe and just refreshes the numbers.

Usage:
  python3 seed_projections_v2.py local   # against local dev (../.env.local)
  python3 seed_projections_v2.py aiven   # against production (.env)
"""

import os
import sys
from decimal import Decimal

import openpyxl
import mysql.connector

SEASON_YEAR = 2026
WEEKS       = 17
SOURCE      = 'internal_model'

# Same alias table as set_2026_prices_v2.py / backfill_rookies.py
NAME_ALIASES = {
    'Cam Ward':              'Cameron Ward',
    'Stetson Bennett':       'Stetson Bennett IV',
    'Joe Milton':            'Joe Milton III',
    'Travis Etienne':        'Travis Etienne Jr.',
    'Aaron Jones':           'Aaron Jones Sr.',
    'Brian Robinson':        'Brian Robinson Jr.',
    'Tyrone Tracy':          'Tyrone Tracy Jr.',
    'Chris Rodriguez':       'Chris Rodriguez Jr.',
    'LeQuint Allen':         'LeQuint Allen Jr.',
    'Ollie Gordon':          'Ollie Gordon II',
    'Michael Pittman':       'Michael Pittman Jr.',
    'D.J. Moore':            'DJ Moore',
    'Luther Burden':         'Luther Burden III',
    'Chris Godwin':          'Chris Godwin Jr.',
    'Tre Harris':            "Tre' Harris",
    'Demario Douglas':       'DeMario Douglas',
    'Theo Wease':            'Theo Wease Jr.',
    'Marvin Mims':           'Marvin Mims Jr.',
    'KeAndre Lambert-Smith': 'Keandre Lambert-Smith',
    'Kevin Austin':          'Kevin Austin Jr.',
    "Dont'e Thornton":       "Dont'e Thornton Jr.",
    'Terrace Marshall':      'Terrace Marshall Jr.',
    'Jimmy Horn':            'Jimmy Horn Jr.',
    'Efton Chism':           'Efton Chism III',
    'Harold Fannin':         'Harold Fannin Jr.',
    'Kyle Pitts':            'Kyle Pitts Sr.',
    'Chigoziem Okonkwo':     'Chig Okonkwo',
    'Erick All':             'Erick All Jr.',
    'Andrew Ogletree':       'Drew Ogletree',
    'Deebo Samuel':          'Deebo Samuel Sr.',
    'James Cook III':        'James Cook',
}


def load_env(path: str) -> dict:
    env = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line and '=' in line and not line.startswith('#'):
                k, _, v = line.partition('=')
                env[k.strip()] = v.strip()
    return env


def db_config(target: str) -> dict:
    here = os.path.dirname(__file__)
    if target == 'aiven':
        env = load_env(os.path.join(here, '.env'))
    elif target == 'local':
        env = load_env(os.path.join(here, '..', '.env.local'))
    else:
        raise SystemExit(f'Unknown target: {target!r} (expected "local" or "aiven")')
    return dict(
        host=env['MYSQL_HOST'],
        port=int(env.get('MYSQL_PORT', 3306)),
        user=env['MYSQL_USER'],
        password=env.get('MYSQL_PASSWORD', ''),
        database=env['MYSQL_DATABASE'],
    )


def main():
    if len(sys.argv) != 2 or sys.argv[1] not in ('local', 'aiven'):
        raise SystemExit('Usage: python3 seed_projections_v2.py [local|aiven]')
    target = sys.argv[1]

    xlsx_path = os.path.join(os.path.dirname(__file__), '..', 'oneleague_prices2.xlsx')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_proj = {}  # db_name -> Decimal season-long PPR projection
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            name, proj = row[0], row[4]
            if not name or proj is None:
                continue
            db_name = NAME_ALIASES.get(name.strip(), name.strip())
            sheet_proj[db_name] = Decimal(str(proj))
    print(f'Loaded {len(sheet_proj)} projections from oneleague_prices2.xlsx')

    conn = mysql.connector.connect(**db_config(target))
    cur = conn.cursor()

    cur.execute("SELECT id, full_name FROM players WHERE position IN ('QB','RB','WR','TE')")
    db_players = cur.fetchall()

    rows = []
    matched_names = set()
    for player_id, full_name in db_players:
        season_proj = sheet_proj.get(full_name)
        if season_proj is None:
            continue
        matched_names.add(full_name)
        weekly = (season_proj / WEEKS).quantize(Decimal('0.01'))
        for week in range(1, WEEKS + 1):
            rows.append((player_id, SEASON_YEAR, week, SOURCE, weekly))

    print(f'Matched {len(matched_names)} players ({len(rows)} weekly rows) — upserting...')

    cur.executemany(
        """
        INSERT INTO player_weekly_projections
            (player_id, season_year, week, projection_source, expected_points)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE expected_points = VALUES(expected_points)
        """,
        rows,
    )
    conn.commit()
    print(f'Done against {target}. {cur.rowcount} rows affected.')

    unmatched_in_sheet = set(sheet_proj.keys()) - matched_names
    if unmatched_in_sheet:
        print(f'\n{len(unmatched_in_sheet)} sheet names had no matching DB player (name mismatch?):')
        for n in sorted(unmatched_in_sheet):
            print(f'  - {n}')

    cur.close()
    conn.close()


if __name__ == '__main__':
    main()
