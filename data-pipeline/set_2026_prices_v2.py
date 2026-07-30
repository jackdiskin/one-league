"""
set_2026_prices_v2.py
──────────────────────
Full price reset from oneleague_prices2.xlsx (QB/RB/WR/TE sheets), matched by
name (same alias table as migrate_to_sportsdata.py — ESPN-ID matching is dead
now that players.external_player_id holds SportsDataIO IDs, not ESPN IDs).

Every QB/RB/WR/TE player gets base_weekly_price == current_price == the
sheet price (rounded to the nearest $0.1M), or the existing $4.5M default
floor if their name isn't in the sheet. intraday_high/low reset to match;
buy/sell counters and net_order_flow reset to 0 (this is a full reset, not
an incremental update). Kickers are left completely untouched — no K sheet
exists, K isn't part of the roster format.

Dumps `players` and this season's `player_market_state` rows to a timestamped
JSON backup before writing anything.

Usage:
  python3 set_2026_prices_v2.py local   # against local dev (../.env.local)
  python3 set_2026_prices_v2.py aiven   # against production (.env)
"""

import json
import os
import sys
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
import mysql.connector

SEASON_YEAR   = 2026
DEFAULT_PRICE = Decimal('4500000')
FANTASY_POS   = {'QB', 'RB', 'WR', 'TE'}

# Spreadsheet name -> DB name aliases (same table as migrate_to_sportsdata.py)
NAME_ALIASES = {
    'Cam Ward':              'Cameron Ward',
    'Stetson Bennett':       'Stetson Bennett IV',
    'Joe Milton':            'Joe Milton III',
    'Travis Etienne':        'Travis Etienne Jr.',
    'Aaron Jones':           'Aaron Jones Sr.',
    'Brian Robinson':        'Brian Robinson Jr.',
    'Tyrone Tracy':          'Tyrone Tracy Jr.',
    'Chris Rodriguez':       'Chris Rodriguez Jr.',
    'LeQuint Allen':         'LeQuint Allen Jr.',
    'Ollie Gordon':          'Ollie Gordon II',
    'Michael Pittman':       'Michael Pittman Jr.',
    'D.J. Moore':            'DJ Moore',
    'Luther Burden':         'Luther Burden III',
    'Chris Godwin':          'Chris Godwin Jr.',
    'Tre Harris':            "Tre' Harris",
    'Demario Douglas':       'DeMario Douglas',
    'Theo Wease':            'Theo Wease Jr.',
    'Marvin Mims':           'Marvin Mims Jr.',
    'KeAndre Lambert-Smith': 'Keandre Lambert-Smith',
    'Kevin Austin':          'Kevin Austin Jr.',
    "Dont'e Thornton":       "Dont'e Thornton Jr.",
    'Terrace Marshall':      'Terrace Marshall Jr.',
    'Jimmy Horn':            'Jimmy Horn Jr.',
    'Efton Chism':           'Efton Chism III',
    'Harold Fannin':         'Harold Fannin Jr.',
    'Kyle Pitts':            'Kyle Pitts Sr.',
    'Chigoziem Okonkwo':     'Chig Okonkwo',
    'Erick All':             'Erick All Jr.',
    'Andrew Ogletree':       'Drew Ogletree',
    'Deebo Samuel':          'Deebo Samuel Sr.',
    'James Cook III':        'James Cook',
}


def round_tenth_million(dollars: Decimal) -> Decimal:
    millions = (dollars / Decimal('1000000')).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
    return millions * Decimal('1000000')


def load_env(path: str) -> dict:
    env = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line and '=' in line and not line.startswith('#'):
                k, _, v = line.partition('=')
                env[k.strip()] = v.strip()
    return env


def db_config(target: str) -> dict:
    here = os.path.dirname(__file__)
    if target == 'aiven':
        env = load_env(os.path.join(here, '.env'))
    elif target == 'local':
        env = load_env(os.path.join(here, '..', '.env.local'))
    else:
        raise SystemExit(f'Unknown target: {target!r} (expected "local" or "aiven")')
    return dict(
        host=env['MYSQL_HOST'],
        port=int(env.get('MYSQL_PORT', 3306)),
        user=env['MYSQL_USER'],
        password=env.get('MYSQL_PASSWORD', ''),
        database=env['MYSQL_DATABASE'],
    )


def main():
    if len(sys.argv) != 2 or sys.argv[1] not in ('local', 'aiven'):
        raise SystemExit('Usage: python3 set_2026_prices_v2.py [local|aiven]')
    target = sys.argv[1]

    # ── Load spreadsheet prices ────────────────────────────────────────────
    xlsx_path = os.path.join(os.path.dirname(__file__), '..', 'oneleague_prices2.xlsx')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_prices = {}  # db_name -> Decimal price in dollars
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            name, price = row[0], row[5]
            if not name or price is None:
                continue
            db_name = NAME_ALIASES.get(name.strip(), name.strip())
            sheet_prices[db_name] = round_tenth_million(Decimal(str(price)) * Decimal('1000000'))
    print(f'Loaded {len(sheet_prices)} prices from oneleague_prices2.xlsx')

    # ── Connect ─────────────────────────────────────────────────────────────
    conn = mysql.connector.connect(**db_config(target))
    cursor = conn.cursor(dictionary=True)

    # ── Backup ──────────────────────────────────────────────────────────────
    cursor.execute('SELECT * FROM players')
    players_backup = cursor.fetchall()
    cursor.execute('SELECT * FROM player_market_state WHERE season_year = %s', (SEASON_YEAR,))
    market_backup = cursor.fetchall()

    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%dT%H%M%S')
    backup_path = os.path.join(backup_dir, f'{target}_pre_price_reset_{stamp}.json')
    with open(backup_path, 'w') as f:
        json.dump({'players': players_backup, 'player_market_state': market_backup}, f, default=str)
    print(f'Backed up {len(players_backup)} players + {len(market_backup)} market_state rows -> {backup_path}')

    # ── Fetch every QB/RB/WR/TE player with 2026 market state ─────────────
    cursor.execute(
        """
        SELECT p.id, p.full_name, p.position, pms.current_price
        FROM players p
        JOIN player_market_state pms ON pms.player_id = p.id AND pms.season_year = %s
        WHERE p.position IN ('QB','RB','WR','TE')
        """,
        (SEASON_YEAR,),
    )
    rows = cursor.fetchall()
    print(f'{len(rows)} QB/RB/WR/TE players to reprice (Kickers left untouched)')

    matched = 0
    defaulted = 0
    unmatched_names = []
    updates = []
    for r in rows:
        price = sheet_prices.get(r['full_name'])
        if price is not None:
            matched += 1
        else:
            price = DEFAULT_PRICE
            defaulted += 1
            unmatched_names.append(r['full_name'])
        updates.append((price, price, price, price, r['id']))

    write_cursor = conn.cursor()
    write_cursor.executemany(
        f"""
        UPDATE player_market_state SET
            base_weekly_price = %s,
            current_price     = %s,
            intraday_high     = %s,
            intraday_low      = %s,
            buy_orders_count  = 0,
            sell_orders_count = 0,
            buy_volume        = 0,
            sell_volume       = 0,
            net_order_flow    = 0,
            last_trade_at     = NULL
        WHERE player_id = %s AND season_year = {SEASON_YEAR}
        """,
        updates,
    )
    conn.commit()

    print(f'\nDone against {target}:')
    print(f'  Matched to spreadsheet: {matched}')
    print(f'  Defaulted to $4.5M:     {defaulted}')
    if unmatched_names:
        print(f'\n  Unmatched names (kept/reset to $4.5M default):')
        for n in sorted(unmatched_names):
            print(f'    - {n}')

    cursor.close()
    write_cursor.close()
    conn.close()


if __name__ == '__main__':
    main()
