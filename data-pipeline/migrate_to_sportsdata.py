"""
migrate_to_sportsdata.py
────────────────────────
Migrates the players table from ESPN-based IDs to SportsDataIO PlayerIDs.

Steps:
  1. Truncate all player-dependent tables (in FK-safe order)
  2. Truncate players
  3. Insert all fantasy-position (QB/RB/WR/TE/K) players from players.json
     - external_player_id = SportsDataIO PlayerID (string)
     - espn_athlete_id left NULL (no longer used)
  4. Assign 2026 initial prices from oneleague_prices1.xlsx via name matching
     (29 name aliases handle Jr./Sr./suffix differences between sources)
  5. Re-populate player_market_state and player_price_weeks for 2026

Usage:
  python3 migrate_to_sportsdata.py
"""

import json
import openpyxl
import mysql.connector
import os
from decimal import Decimal

# ── Config ────────────────────────────────────────────────────────────────────
env_path = os.path.join(os.path.dirname(__file__), '.env')
env = {}
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if line and '=' in line and not line.startswith('#'):
            k, _, v = line.partition('=')
            env[k.strip()] = v.strip()

DB = dict(
    host=env['MYSQL_HOST'],
    port=int(env.get('MYSQL_PORT', 3306)),
    user=env['MYSQL_USER'],
    password=env['MYSQL_PASSWORD'],
    database=env['MYSQL_DATABASE'],
)

SEASON_YEAR   = 2026
WEEK          = 1
DEFAULT_PRICE = Decimal('4_500_000')
FANTASY_POS   = {'QB', 'RB', 'WR', 'TE', 'K'}

# Spreadsheet name -> SportsDataIO name aliases (all 29 mismatches)
NAME_ALIASES = {
    'Cam Ward':              'Cameron Ward',
    'Stetson Bennett':       'Stetson Bennett IV',
    'Joe Milton':            'Joe Milton III',
    'Travis Etienne':        'Travis Etienne Jr.',
    'Aaron Jones':           'Aaron Jones Sr.',
    'Brian Robinson':        'Brian Robinson Jr.',
    'Tyrone Tracy':          'Tyrone Tracy Jr.',
    'Chris Rodriguez':       'Chris Rodriguez Jr.',
    'LeQuint Allen':         'LeQuint Allen Jr.',
    'Ollie Gordon':          'Ollie Gordon II',
    'Michael Pittman':       'Michael Pittman Jr.',
    'D.J. Moore':            'DJ Moore',
    'Luther Burden':         'Luther Burden III',
    'Chris Godwin':          'Chris Godwin Jr.',
    'Tre Harris':            "Tre' Harris",
    'Demario Douglas':       'DeMario Douglas',
    'Theo Wease':            'Theo Wease Jr.',
    'Marvin Mims':           'Marvin Mims Jr.',
    'KeAndre Lambert-Smith': 'Keandre Lambert-Smith',
    'Kevin Austin':          'Kevin Austin Jr.',
    "Dont'e Thornton":       "Dont'e Thornton Jr.",
    'Terrace Marshall':      'Terrace Marshall Jr.',
    'Jimmy Horn':            'Jimmy Horn Jr.',
    'Efton Chism':           'Efton Chism III',
    'Harold Fannin':         'Harold Fannin Jr.',
    'Kyle Pitts':            'Kyle Pitts Sr.',
    'Chigoziem Okonkwo':     'Chig Okonkwo',
    'Erick All':             'Erick All Jr.',
    'Andrew Ogletree':       'Drew Ogletree',
}

# SportsDataIO Status -> DB status enum
STATUS_MAP = {
    'Active':          'active',
    'Injured Reserve': 'injured',
    'Suspended':       'inactive',
    'Inactive':        'inactive',
}

# Tables to clear in FK-safe order (players last)
CLEAR_TABLES = [
    'player_price_ticks',
    'fantasy_team_roster',
    'fantasy_team_weekly_scores',
    'player_transactions',
    'fantasy_teams',
    'player_weekly_projections',
    'player_weekly_scores',
    'player_price_weeks',
    'player_market_state',
    'players',
]

# ── Step 1: Load spreadsheet prices (keyed by SportsDataIO name) ──────────────
print('Loading spreadsheet prices...')
xlsx = os.path.join(os.path.dirname(__file__), 'oneleague_prices1.xlsx')
wb   = openpyxl.load_workbook(xlsx, data_only=True)

sheet_prices = {}  # sdio_name -> Decimal price in dollars
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name, _, _, _, _, price = row[0], row[1], row[2], row[3], row[4], row[5]
        if not name or price is None:
            continue
        sdio_name = NAME_ALIASES.get(name.strip(), name.strip())
        sheet_prices[sdio_name] = Decimal(str(price)) * Decimal('1_000_000')

print(f'  {len(sheet_prices)} price entries loaded (including aliases)')

# ── Step 2: Load SportsDataIO players ────────────────────────────────────────
print('\nLoading SportsDataIO players...')
json_path = os.path.join(os.path.dirname(__file__), 'players.json')
with open(json_path) as f:
    raw = json.load(f)

players = [p for p in raw if p.get('FantasyPosition') in FANTASY_POS]
print(f'  {len(players)} fantasy-position players (QB/RB/WR/TE/K)')
print(f'  {sum(1 for p in players if p.get("Active"))} active, '
      f'{sum(1 for p in players if not p.get("Active"))} inactive/IR')

# ── Step 3: Preview + confirm ─────────────────────────────────────────────────
priced   = sum(1 for p in players if p['Name'] in sheet_prices)
defaulted = len(players) - priced
print(f'\nPrice assignment preview:')
print(f'  Matched to spreadsheet: {priced}')
print(f'  Will use default ${DEFAULT_PRICE/1_000_000}M: {defaulted}')
print()
print(f'Tables to be TRUNCATED (all data lost):')
for t in CLEAR_TABLES:
    print(f'  {t}')
print()
resp = input('Type YES to proceed: ').strip()
if resp != 'YES':
    print('Aborting.')
    exit(0)

# ── Step 4: Connect and truncate ──────────────────────────────────────────────
conn   = mysql.connector.connect(**DB)
cursor = conn.cursor()

print('\nTruncating tables...')
cursor.execute('SET FOREIGN_KEY_CHECKS=0')
for table in CLEAR_TABLES:
    cursor.execute(f'DELETE FROM `{table}`')
    print(f'  Cleared {table} ({cursor.rowcount} rows)')
cursor.execute('SET FOREIGN_KEY_CHECKS=1')
conn.commit()

# ── Step 5: Insert players ────────────────────────────────────────────────────
print('\nInserting SportsDataIO players...')

player_rows = []
for p in players:
    sdio_status = p.get('Status') or 'Inactive'
    db_status   = STATUS_MAP.get(sdio_status, 'inactive')
    team_code   = p.get('Team') or 'FA'
    headshot    = p.get('UsaTodayHeadshotNoBackgroundUrl') or p.get('UsaTodayHeadshotUrl')

    player_rows.append((
        str(p['PlayerID']),         # external_player_id
        p['Name'],                  # full_name
        p.get('ShortName'),         # short_name
        team_code,                  # team_code
        team_code,                  # team_name (code only; no name mapping available)
        p['FantasyPosition'],       # position
        db_status,                  # status
        headshot,                   # headshot_url
        p.get('Number'),            # jersey_number
    ))

cursor.executemany("""
    INSERT INTO players
        (external_player_id, full_name, short_name, team_code, team_name,
         position, status, headshot_url, jersey_number)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
""", player_rows)
conn.commit()
print(f'  Inserted {cursor.rowcount} players')

# ── Step 6: Fetch inserted player IDs ─────────────────────────────────────────
cursor.execute('SELECT id, full_name FROM players')
db_players = cursor.fetchall()  # list of (id, full_name)
print(f'  Verified {len(db_players)} players in DB')

# ── Step 7: Insert 2026 market state + price weeks ───────────────────────────
print('\nSeeding 2026 prices...')

market_rows = []
week_rows   = []
matched = defaulted_n = 0

for player_id, full_name in db_players:
    price = sheet_prices.get(full_name, DEFAULT_PRICE)
    if full_name in sheet_prices:
        matched += 1
    else:
        defaulted_n += 1

    market_rows.append((player_id, SEASON_YEAR, WEEK, price, price, price, price))
    week_rows.append((player_id, SEASON_YEAR, WEEK, price, price, price))

cursor.executemany("""
    INSERT INTO player_market_state
        (player_id, season_year, current_week,
         base_weekly_price, current_price, intraday_high, intraday_low)
    VALUES (%s, %s, %s, %s, %s, %s, %s)
""", market_rows)

cursor.executemany("""
    INSERT INTO player_price_weeks
        (player_id, season_year, week, opening_price, base_price, closing_price)
    VALUES (%s, %s, %s, %s, %s, %s)
""", week_rows)

conn.commit()
print(f'  Priced from spreadsheet: {matched}')
print(f'  Priced at default ${DEFAULT_PRICE/1_000_000}M: {defaulted_n}')

# ── Step 8: Verify ────────────────────────────────────────────────────────────
cursor.execute('SELECT COUNT(*) FROM players')
print(f'\nVerification:')
print(f'  players:              {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM player_market_state WHERE season_year=%s', (SEASON_YEAR,))
print(f'  player_market_state:  {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM player_price_weeks WHERE season_year=%s AND week=%s', (SEASON_YEAR, WEEK))
print(f'  player_price_weeks:   {cursor.fetchone()[0]}')

cursor.close()
conn.close()
print('\nDone.')
