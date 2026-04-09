"""
Set initial prices for the 2026 season.
- Uses oneleague_prices1.xlsx ESPN ID -> Price mapping (all sheets: QB, RB, WR, TEs)
- Players with ESPN IDs found in spreadsheet get that price
- Players with ESPN IDs NOT in spreadsheet get default 4.5
- Players without ESPN IDs get default 4.5
Inserts rows into player_market_state and player_price_weeks (week=1) for season_year=2026.
"""

import openpyxl
import mysql.connector
import os
from decimal import Decimal

# Load env from .env file
env_path = os.path.join(os.path.dirname(__file__), '.env')
env = {}
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if line and '=' in line and not line.startswith('#'):
            key, _, val = line.partition('=')
            env[key.strip()] = val.strip()

DB_CONFIG = {
    'host': env['MYSQL_HOST'],
    'port': int(env.get('MYSQL_PORT', 3306)),
    'user': env['MYSQL_USER'],
    'password': env['MYSQL_PASSWORD'],
    'database': env['MYSQL_DATABASE'],
}

DEFAULT_PRICE = Decimal('4500000')
MILLION = Decimal('1000000')
SEASON_YEAR = 2026
WEEK = 1

# --- Step 1: Load spreadsheet prices ---
xlsx_path = os.path.join(os.path.dirname(__file__), 'oneleague_prices1.xlsx')
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

espn_prices = {}  # espn_id (str) -> price (Decimal)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            header_found = True
            continue  # skip header row
        # row: (Name, Team, Position, ESPN ID, Projected Pts, Price ($M))
        if len(row) < 6:
            continue
        espn_id = row[3]
        price = row[5]
        if espn_id is not None and price is not None:
            try:
                espn_prices[str(int(espn_id))] = Decimal(str(price)) * MILLION
            except (ValueError, TypeError):
                pass

print(f"Loaded {len(espn_prices)} ESPN ID -> price mappings from spreadsheet")

# Show a few examples
sample = list(espn_prices.items())[:5]
for k, v in sample:
    print(f"  ESPN {k} -> ${v:,.0f}")

# --- Step 2: Fetch all players from DB ---
conn = mysql.connector.connect(**DB_CONFIG)
cursor = conn.cursor(dictionary=True)

cursor.execute("SELECT id, full_name, espn_athlete_id FROM players")
players = cursor.fetchall()
print(f"\nFetched {len(players)} players from DB")

# --- Step 3: Check for existing 2026 market state rows ---
cursor.execute("SELECT COUNT(*) as cnt FROM player_market_state WHERE season_year = %s", (SEASON_YEAR,))
existing = cursor.fetchone()['cnt']
if existing > 0:
    print(f"\nWARNING: {existing} rows already exist in player_market_state for {SEASON_YEAR}.")
    resp = input("Delete existing 2026 rows and re-insert? (yes/no): ").strip().lower()
    if resp == 'yes':
        cursor.execute("DELETE FROM player_market_state WHERE season_year = %s", (SEASON_YEAR,))
        cursor.execute("DELETE FROM player_price_weeks WHERE season_year = %s", (SEASON_YEAR,))
        conn.commit()
        print("Deleted existing 2026 rows.")
    else:
        print("Aborting.")
        cursor.close()
        conn.close()
        exit(0)

# --- Step 4: Build price rows ---
market_state_rows = []
price_week_rows = []

matched = 0
defaulted_espn = 0
defaulted_no_espn = 0

for player in players:
    player_id = player['id']
    espn_id = player['espn_athlete_id']

    if espn_id and espn_id in espn_prices:
        price = espn_prices[espn_id]
        matched += 1
    else:
        price = DEFAULT_PRICE
        if espn_id:
            defaulted_espn += 1
        else:
            defaulted_no_espn += 1

    market_state_rows.append((
        player_id, SEASON_YEAR, WEEK,
        price,   # base_weekly_price
        price,   # current_price
        price,   # intraday_high
        price,   # intraday_low
    ))
    price_week_rows.append((
        player_id, SEASON_YEAR, WEEK,
        price,   # opening_price
        price,   # base_price
        price,   # closing_price
    ))

print(f"\nPrice assignment summary:")
print(f"  Matched from spreadsheet: {matched}")
print(f"  Has ESPN ID but not in spreadsheet (default {DEFAULT_PRICE}): {defaulted_espn}")
print(f"  No ESPN ID (default {DEFAULT_PRICE}): {defaulted_no_espn}")
print(f"  Total rows to insert: {len(market_state_rows)}")

# --- Step 5: Insert ---
print(f"\nInserting {len(market_state_rows)} rows into player_market_state...")
cursor.executemany("""
    INSERT INTO player_market_state
        (player_id, season_year, current_week, base_weekly_price, current_price, intraday_high, intraday_low)
    VALUES (%s, %s, %s, %s, %s, %s, %s)
""", market_state_rows)

print(f"Inserting {len(price_week_rows)} rows into player_price_weeks...")
cursor.executemany("""
    INSERT INTO player_price_weeks
        (player_id, season_year, week, opening_price, base_price, closing_price)
    VALUES (%s, %s, %s, %s, %s, %s)
""", price_week_rows)

conn.commit()
print("Done! Committed all rows.")

# --- Step 6: Verify ---
cursor.execute("SELECT COUNT(*) as cnt FROM player_market_state WHERE season_year = %s", (SEASON_YEAR,))
ms_count = cursor.fetchone()['cnt']
cursor.execute("SELECT COUNT(*) as cnt FROM player_price_weeks WHERE season_year = %s AND week = %s", (SEASON_YEAR, WEEK))
pw_count = cursor.fetchone()['cnt']
print(f"\nVerification:")
print(f"  player_market_state 2026: {ms_count} rows")
print(f"  player_price_weeks 2026 week 1: {pw_count} rows")

cursor.close()
conn.close()
